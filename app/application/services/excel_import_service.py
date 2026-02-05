"""
Service za import formi iz Excel fajlova.

Podržava:
- Generisanje Excel template-a
- Parsiranje i validaciju Excel fajlova
- Konverziju u FormSchemaCreate format
"""
from io import BytesIO
from typing import List, Tuple, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.api.form_schema import (
    FormSchemaCreate,
    FormField,
    FieldType,
    FieldOption
)


# Dozvoljeni tipovi polja
ALLOWED_FIELD_TYPES = {
    "text": FieldType.TEXT,
    "email": FieldType.EMAIL,
    "number": FieldType.NUMBER,
    "textarea": FieldType.TEXTAREA,
    "select": FieldType.SELECT,
    "radio": FieldType.RADIO,
    "checkbox": FieldType.CHECKBOX,
    "date": FieldType.DATE,
}

# Tipovi koji zahtevaju options
TYPES_REQUIRING_OPTIONS = {"select", "radio"}

# Excel styling
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelImportService:
    """Service za Excel import operacije."""
    
    # =========================================================================
    # Template Generation
    # =========================================================================
    
    def generate_template(self) -> BytesIO:
        """
        Generiše prazan Excel template za import formi.
        
        Returns:
            BytesIO: Excel fajl kao byte stream
        """
        wb = Workbook()
        
        # === Sheet 1: Form ===
        ws_form = wb.active
        ws_form.title = "Form"
        
        form_headers = ["name", "description"]
        form_example = ["Moja Forma", "Opis forme (opciono)"]
        
        self._write_header_row(ws_form, form_headers)
        self._write_data_row(ws_form, 2, form_example)
        self._auto_column_width(ws_form, form_headers)
        
        # === Sheet 2: Fields ===
        ws_fields = wb.create_sheet("Fields")
        
        field_headers = ["id", "label", "type", "placeholder", "required", "options", "min", "max"]
        field_examples = [
            ["name", "Ime i Prezime", "text", "Unesite ime...", "TRUE", "", "", ""],
            ["email", "Email Adresa", "email", "vas@email.com", "TRUE", "", "", ""],
            ["age", "Godine", "number", "", "FALSE", "", "18", "99"],
            ["country", "Država", "select", "Izaberite državu", "TRUE", "Srbija, Hrvatska, BiH", "", ""],
            ["gender", "Pol", "radio", "", "FALSE", "Muški, Ženski, Drugo", "", ""],
            ["newsletter", "Želim newsletter", "checkbox", "", "FALSE", "", "", ""],
            ["birth_date", "Datum rođenja", "date", "", "FALSE", "", "", ""],
            ["message", "Poruka", "textarea", "Unesite poruku...", "FALSE", "", "", ""],
        ]
        
        self._write_header_row(ws_fields, field_headers)
        for i, row in enumerate(field_examples, start=2):
            self._write_data_row(ws_fields, i, row)
        self._auto_column_width(ws_fields, field_headers)
        
        # === Sheet 3: Instructions ===
        ws_help = wb.create_sheet("Instructions")
        instructions = [
            ["FormForge - Excel Import Template"],
            [""],
            ["SHEET: Form"],
            ["- name: Naziv forme (obavezno)"],
            ["- description: Opis forme (opciono)"],
            [""],
            ["SHEET: Fields"],
            ["- id: Jedinstveni ID polja (obavezno, bez razmaka)"],
            ["- label: Labela koja se prikazuje korisniku (obavezno)"],
            ["- type: Tip polja (obavezno)"],
            ["  Dozvoljeni tipovi: text, email, number, textarea, select, radio, checkbox, date"],
            ["- placeholder: Placeholder tekst (opciono)"],
            ["- required: TRUE ili FALSE (opciono, default: FALSE)"],
            ["- options: Za select/radio - opcije razdvojene zarezom (obavezno za select/radio)"],
            ["- min: Minimalna vrednost za number polja (obavezno za number)"],
            ["- max: Maksimalna vrednost za number polja (obavezno za number)"],
            [""],
            ["NAPOMENE:"],
            ["- Ne menjajte nazive sheet-ova (Form, Fields)"],
            ["- Ne menjajte header redove"],
            ["- Možete obrisati primer redove i dodati svoje"],
            ["- ID polja mora biti jedinstven"],
        ]
        
        for i, row in enumerate(instructions, start=1):
            ws_help.cell(row=i, column=1, value=row[0])
            if i == 1:
                ws_help.cell(row=i, column=1).font = Font(bold=True, size=14)
        
        ws_help.column_dimensions['A'].width = 80
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    # =========================================================================
    # Excel Parsing & Validation
    # =========================================================================
    
    def parse_excel(self, file_content: bytes) -> Tuple[bool, Optional[FormSchemaCreate], List[str]]:
        """
        Parsira Excel fajl i vraća FormSchemaCreate ili listu grešaka.
        
        Args:
            file_content: Sadržaj Excel fajla kao bytes
        
        Returns:
            Tuple[bool, Optional[FormSchemaCreate], List[str]]:
                - success: Da li je parsiranje uspelo
                - form: FormSchemaCreate ako je uspelo, None ako nije
                - errors: Lista grešaka (prazna ako je uspelo)
        """
        errors = []
        
        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        except Exception as e:
            return False, None, [f"Invalid Excel file: {str(e)}"]
        
        # === Validacija Sheet-ova ===
        if "Form" not in wb.sheetnames:
            errors.append("Missing required sheet: 'Form'")
        if "Fields" not in wb.sheetnames:
            errors.append("Missing required sheet: 'Fields'")
        
        if errors:
            return False, None, errors
        
        # === Parsiranje Form Sheet-a ===
        ws_form = wb["Form"]
        form_data, form_errors = self._parse_form_sheet(ws_form)
        errors.extend(form_errors)
        
        # === Parsiranje Fields Sheet-a ===
        ws_fields = wb["Fields"]
        fields, field_errors = self._parse_fields_sheet(ws_fields)
        errors.extend(field_errors)
        
        if errors:
            return False, None, errors
        
        # === Kreiranje FormSchemaCreate ===
        form_schema = FormSchemaCreate(
            name=form_data["name"],
            description=form_data.get("description"),
            fields=fields,
            rules=None,
            theme=None
        )
        
        return True, form_schema, []
    
    def _parse_form_sheet(self, ws) -> Tuple[dict, List[str]]:
        """Parsira Form sheet."""
        errors = []
        form_data = {}
        
        # Očekivani headeri
        headers = self._get_row_values(ws, 1)
        
        if not headers or "name" not in [h.lower() if h else "" for h in headers]:
            errors.append("Form sheet: Missing 'name' column")
            return form_data, errors
        
        # Mapiranje headera na indekse
        header_map = {h.lower(): i for i, h in enumerate(headers) if h}
        
        # Čitanje podataka iz reda 2
        values = self._get_row_values(ws, 2)
        
        if not values or len(values) == 0:
            errors.append("Form sheet: No data found in row 2")
            return form_data, errors
        
        # Name
        name_idx = header_map.get("name")
        if name_idx is not None and name_idx < len(values):
            name = values[name_idx]
            if name and str(name).strip():
                form_data["name"] = str(name).strip()
            else:
                errors.append("Form name is required")
        else:
            errors.append("Form name is required")
        
        # Description (opciono)
        desc_idx = header_map.get("description")
        if desc_idx is not None and desc_idx < len(values):
            desc = values[desc_idx]
            if desc and str(desc).strip():
                form_data["description"] = str(desc).strip()
        
        return form_data, errors
    
    def _parse_fields_sheet(self, ws) -> Tuple[List[FormField], List[str]]:
        """Parsira Fields sheet."""
        errors = []
        fields = []
        seen_ids = set()
        
        # Očekivani headeri
        headers = self._get_row_values(ws, 1)
        
        if not headers:
            errors.append("Fields sheet: No headers found")
            return fields, errors
        
        # Mapiranje headera na indekse (case-insensitive)
        header_map = {str(h).lower().strip(): i for i, h in enumerate(headers) if h}
        
        # Provera obaveznih kolona
        required_columns = ["id", "label", "type"]
        for col in required_columns:
            if col not in header_map:
                errors.append(f"Fields sheet: Missing required column '{col}'")
        
        if errors:
            return fields, errors
        
        # Parsiranje redova (od reda 2)
        row_num = 2
        while True:
            values = self._get_row_values(ws, row_num)
            
            # Prekini ako je red prazan
            if not values or all(v is None or str(v).strip() == "" for v in values):
                break
            
            field, field_errors = self._parse_field_row(values, header_map, row_num, seen_ids)
            
            if field_errors:
                errors.extend(field_errors)
            elif field:
                fields.append(field)
                seen_ids.add(field.id)
            
            row_num += 1
        
        # Provera da ima bar jedno polje
        if not fields and not errors:
            errors.append("At least one field is required")
        
        return fields, errors
    
    def _parse_field_row(
        self, 
        values: List, 
        header_map: dict, 
        row_num: int,
        seen_ids: set
    ) -> Tuple[Optional[FormField], List[str]]:
        """Parsira jedan red iz Fields sheet-a."""
        errors = []
        
        def get_value(column: str) -> Optional[str]:
            idx = header_map.get(column)
            if idx is not None and idx < len(values):
                val = values[idx]
                if val is not None:
                    return str(val).strip()
            return None
        
        # === ID ===
        field_id = get_value("id")
        if not field_id:
            errors.append(f"Row {row_num}: Field ID is required")
            return None, errors
        
        # Validacija ID formata (bez razmaka, samo alfanumerik i _)
        if not field_id.replace("_", "").isalnum():
            errors.append(f"Row {row_num}: Field ID '{field_id}' contains invalid characters (use only letters, numbers, underscore)")
        
        # Duplikat check
        if field_id in seen_ids:
            errors.append(f"Row {row_num}: Duplicate field ID: '{field_id}'")
            return None, errors
        
        # === Label ===
        label = get_value("label")
        if not label:
            errors.append(f"Row {row_num}: Field label is required")
            return None, errors
        
        # === Type ===
        field_type_str = get_value("type")
        if not field_type_str:
            errors.append(f"Row {row_num}: Field type is required")
            return None, errors
        
        field_type_str = field_type_str.lower()
        if field_type_str not in ALLOWED_FIELD_TYPES:
            allowed = ", ".join(ALLOWED_FIELD_TYPES.keys())
            errors.append(f"Row {row_num}: Invalid field type '{field_type_str}'. Allowed: {allowed}")
            return None, errors
        
        field_type = ALLOWED_FIELD_TYPES[field_type_str]
        
        # === Placeholder ===
        placeholder = get_value("placeholder")
        
        # === Required ===
        required_str = get_value("required")
        required = False
        if required_str:
            required = required_str.upper() in ("TRUE", "YES", "1", "DA")
        
        # === Options (za select/radio) ===
        options_str = get_value("options")
        options = []
        
        if field_type_str in TYPES_REQUIRING_OPTIONS:
            if not options_str:
                errors.append(f"Row {row_num}: Field '{field_id}' of type '{field_type_str}' requires options")
            else:
                options = self._parse_options(options_str)
                if not options:
                    errors.append(f"Row {row_num}: Field '{field_id}' has invalid options format")
        elif options_str:
            # Ako ima options ali nije select/radio, svejedno ih parsiramo (za checkbox npr)
            options = self._parse_options(options_str)
        
        # === Min/Max (za number) ===
        min_val = get_value("min")
        max_val = get_value("max")
        validations = []
        
        if field_type_str == "number":
            if not min_val:
                errors.append(f"Row {row_num}: Field '{field_id}' of type 'number' requires 'min' value")
            if not max_val:
                errors.append(f"Row {row_num}: Field '{field_id}' of type 'number' requires 'max' value")
            
            if min_val and max_val:
                try:
                    min_num = float(min_val)
                    max_num = float(max_val)
                    
                    if min_num > max_num:
                        errors.append(f"Row {row_num}: Field '{field_id}' has min ({min_val}) greater than max ({max_val})")
                    else:
                        validations.append({"type": "min", "value": min_num})
                        validations.append({"type": "max", "value": max_num})
                except ValueError:
                    errors.append(f"Row {row_num}: Field '{field_id}' has invalid min/max values (must be numbers)")
        
        # Required validacija
        if required:
            validations.append({"type": "required", "value": True})
        
        if errors:
            return None, errors
        
        # Kreiranje FormField
        field = FormField(
            id=field_id,
            type=field_type,
            label=label,
            placeholder=placeholder,
            options=options if options else [],
            validations=validations,
            rules=[]
        )
        
        return field, []
    
    def _parse_options(self, options_str: str) -> List[FieldOption]:
        """Parsira options string u listu FieldOption."""
        options = []
        parts = options_str.split(",")
        
        for part in parts:
            part = part.strip()
            if part:
                options.append(FieldOption(label=part, value=part))
        
        return options
    
    # =========================================================================
    # Helper Methods
    # =========================================================================
    
    def _write_header_row(self, ws, headers: List[str]):
        """Piše header red sa stilom."""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
    
    def _write_data_row(self, ws, row: int, values: List):
        """Piše data red."""
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
    
    def _auto_column_width(self, ws, headers: List[str]):
        """Automatski podešava širinu kolona."""
        for i, header in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max(15, len(header) + 5)
    
    def _get_row_values(self, ws, row: int) -> List:
        """Dohvata vrednosti iz reda."""
        values = []
        for cell in ws[row]:
            values.append(cell.value)
        return values
